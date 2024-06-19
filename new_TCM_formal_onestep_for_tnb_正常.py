import time,os
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.keys import Keys
import openpyxl
import datetime

class TCM:
    #-----定义一些常用参数变量--------
    def __init__(self):
        self.driver=webdriver.Chrome()
        self.excel=openpyxl.load_workbook(r'C:\excel_to_TCM_face_to _all_newTCM.xlsx')
        self.basicinfo_sheet = self.excel['basicinfo']
        self.testplan_sheet = self.excel['testplan']
        self.user_name=self.basicinfo_sheet['B1'].value
        self.password = self.basicinfo_sheet['B2'].value
        self.project_name=self.basicinfo_sheet['B3'].value
        self.testplan_sheet_max_rows=self.testplan_sheet.max_row
    # -----定义一些常用参数变量--------


    #筛选Project,选择not start
    def enter_not_start_case(self):
        self.driver.get("https://tms.wistron.com/#/")
        self.driver.maximize_window()
        login = WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.XPATH, "//a[@class='sc-kEqXSa bAVzgZ']")))
        login.click()
        input_username = WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.ID, "i0116")))
        input_username.send_keys(self.user_name)
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//input[@id='idSIButton9' and @value='Next']").click()
        input_password = WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.ID, "i0118")))
        input_password.send_keys(self.password)
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//input[@id='idSIButton9' and @value='Sign in']").click()
        QT = WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.LINK_TEXT, 'QT')))
        QT.click()
        time.sleep(3)
        not_start = WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.ID, "myTask_testRun_taskCategory_infoListColumnsContent_Not Started")))
        not_start.click()#选择not start的
        time.sleep(5)
        Proect_name_search= WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.ID, "rc_select_1")))
        Proect_name_search.click()
        Proect_name_search.send_keys(self.project_name)
        Proect_name_search.send_keys(Keys.ENTER)
        # level=WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.XPATH, "/html/body/div[1]/section/section/section[2]/div[2]/div[2]/div/div[2]/div[1]/div/div[2]/div/div/div/div/div/div/input")))
        # level.click()
        # level.send_keys('Ma')
        # time.sleep(1)
        # level.send_keys(Keys.ENTER)
        time.sleep(2)
    #筛选Project,选择not start

    def enter_case(self):
        list_1 = ['01', '02', '03', '04', '05', '06', '07', '08', '09']  # 为了后面选开始时间定义一个列表
        main_window = self.driver.current_window_handle
        for every_row in range(2,self.testplan_sheet.max_row + 1):  # 这里是匹配testplan sheet的最大行数，除去第一行表头，+1是因为range(2,9）行是循环2到8行.所以加1刚刚好
            every_row_list = []  # 给每一行建立一个空列表来装每行一具体的值
            for cell in self.testplan_sheet[every_row]:  # testplan_sheet[i] i是就代表testplan sheet的第几行，cell是代表这一行中的每一个单元格
                every_row_list.append(cell.value)
            every_row_list[2] = every_row_list[2].strftime("%Y-%m-%d")  # 开始日期转化为字符串格式，因为从excel读的值datetime.datetime格式
            every_row_list[3] = every_row_list[3].strftime("%Y-%m-%d")  # 结束日期转化为字符串格式，因为从excel读的值datetime.datetime格式
            #因为get plan程序随机生成开始结束日期，所以是datetime格式，所以要转字符串, 生成的开始结束时间在get plan里已经转为了字符串格式，所以这里不需要再转了
            if every_row_list[10]==None:#代表没有issue，执行pass
                search_case=self.driver.find_element(By.ID, "rc_select_10")#找到搜查case的输入框
                search_case.click()
                search_case.send_keys(Keys.CONTROL,'a')#全选ctrl+A
                search_case.send_keys(Keys.BACKSPACE)
                time.sleep(1)
                search_case.send_keys(every_row_list[0])#输入caseID
                search_case.send_keys(Keys.ENTER)
                self.driver.find_element(By.XPATH, "/html/body/div[1]/section/section/section/div/div[2]/div[1]/div/div[2]/div/div[16]/div[2]/div/button[2]/span").click()#点击搜索
                time.sleep(3)#这里3秒是为了等待显示搜索后的页面，否则后面的代码会找到所有的test cases
                try:#我就默认搜索case后至少有1支case出现
                    all_cases=self.driver.find_elements(By.XPATH, "//*[@id='root']/section/section/section/div/div[2]/div[3]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr")#找到搜索后出现的所有case，这里第一个tr是头标，需要去除
                    len_all_cases=len(all_cases)
                    first_case= self.driver.find_element(By.XPATH,"/html/body/div[1]/section/section/section/div/div[2]/div[3]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr[2]/td[3]/div")#锁定正常情况下搜到case的情况，至少有一支case的情况，第4个td标签case ID，如果这里报错，代表搜到了case,但是没有not run的，plan没拉或者被打过
                    if first_case.text==every_row_list[0]:#如果到了这步，代表上面一步pass了，说明有搜到case并至少出现一份或者直接没有搜到这份case，依然会有上一步的标签, 就是默认没搜索前的第一份case
                    #如果相等代表是搜到的情况
                        for every_case in range(2,len_all_cases+1):#，遍历所有case,除去第一个tr，因为是标头，所以从2开始，+1是因为range(2,all_cases）只能取到2到all_cases-1,不能遍历所有搜索到的case
                            every_case = self.driver.find_element(By.XPATH,f"/html/body/div[1]/section/section/section/div/div[2]/div[3]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr[{every_case}]/td[18]")  # 定位case
                            self.driver.execute_script("arguments[0].scrollIntoView();", every_case)  # 移到case 元素可见
                            every_case.click()  # 点击执行case
                            popup_windows = self.driver.window_handles[-1]  # 跳转到新页面执行准备case
                            self.driver.switch_to.window(popup_windows)  # 跳转到新页面执行准备case
                            edit_button=WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.XPATH,"/html/body/div/section/section/div[1]/div[2]/button[1]/span")))
                            edit_button.click()
                            pass_button=WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.XPATH,"/html/body/div/section/section/section[3]/div/div[1]/section/div/div/div/div[1]/div/label[2]/span[1]/input")))
                            self.driver.execute_script("arguments[0].click();",pass_button)
                            config=self.driver.find_element(By.XPATH, "//*[@id='root']/section/section/section[2]/div/div/div/div/div/div/div/div/div/table/tbody/tr/td[6]/div")#获取config 标签
                            config_text=config.text#获取config 标签的text值
                            self.driver.find_element(By.XPATH,"/html/body/div/section/section/section[3]/div/div[1]/div[4]/section[1]/div/div[2]/div/input").send_keys(config_text)#Actual DUT标签填值
                            start_date=self.driver.find_element(By.XPATH,"/html/body/div/section/section/section[3]/div/div[1]/div[4]/section[1]/div/div[3]/div/div/div[1]/div/input")#找到开始日期
                            start_date.send_keys(every_row_list[2])#填写开始日期
                            start_date.send_keys(Keys.ENTER)#按下回车键
                            start_time=self.driver.find_element(By.XPATH,"/html/body/div[1]/section/section/section[3]/div/div[1]/div[4]/section[1]/div/div[3]/div/div/div[2]/div/input")#找到开始时间
                            start_time.send_keys(every_row_list[4])#填写开始时间
                            start_time_hr = every_row_list[4][0:2]  # 取开始时间的前2位小时部分
                            start_time_minute = every_row_list[4][3:5]  # 取开始时间的后2位分钟部分
                            if start_time_hr in list_1:
                                start_time_hr_1 = int(start_time_hr[1]) + 1  # 例如02就取2然后变为整数然后加1就是标签
                                select_time_hr = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[3]/div/div/div/div/div[1]/div/ul[1]/li[{start_time_hr_1}]/div")))  # 锁定开始时间小时标签
                                self.driver.execute_script("arguments[0].click();", select_time_hr)
                            else:
                                start_time_hr_1 = int(start_time_hr) + 1  # 只要不是01-09直接加1就行了，分析标签得出的结论
                                select_time_hr = WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[3]/div/div/div/div/div[1]/div/ul[1]/li[{start_time_hr_1}]/div")))  # 锁定开始时间小时标签
                                self.driver.execute_script("arguments[0].click();", select_time_hr)
                            if start_time_minute in list_1:
                                start_time_minute_1 = int(start_time_minute[1]) + 1  # 例如02就取2然后变为整数然后加1就是标签
                                select_time_minute = WebDriverWait(self.driver, 100).until(
                                    ec.presence_of_element_located((By.XPATH, f"/html/body/div[3]/div/div/div/div/div[1]/div/ul[2]/li[{start_time_minute_1}]/div")))  # 锁定开始时间分钟标签 self.driver.execute_script("arguments[0].click();", select_time_minute)
                            else:
                                start_time_minute_1 = int(start_time_minute) + 1  # 只要不是01-09直接加1就行了，分析标签得出的结论
                                select_time_minute = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[3]/div/div/div/div/div[1]/div/ul[2]/li[{start_time_minute_1}]/div")))  # 锁定开始时间分钟标签
                                self.driver.execute_script("arguments[0].click();", select_time_minute)
                            start_time.send_keys(Keys.ENTER)
                            end_date=self.driver.find_element(By.XPATH,"/html/body/div[1]/section/section/section[3]/div/div[1]/div[4]/section[1]/div/div[4]/div/div/div[1]/div/input")#找结束日期
                            end_date.click()
                            end_date.send_keys(Keys.CONTROL, 'a')
                            end_date.send_keys(Keys.BACKSPACE)  # 填写开始日期后默认日期默认填了开始日期，所有先去除默认日期
                            end_date.send_keys(every_row_list[3])#填写结束日期
                            end_date.send_keys(Keys.ENTER)
                            end_time=self.driver.find_element(By.XPATH,"/html/body/div[1]/section/section/section[3]/div/div[1]/div[4]/section[1]/div/div[4]/div/div/div[2]/div/input")#找结束时间
                            end_time.send_keys(every_row_list[5])#填写结束时间
                            end_time_hr = every_row_list[5][0:2]  # 取结束时间的前2位小时部分
                            end_time_minute = every_row_list[5][3:5]
                            if end_time_hr in list_1:
                                end_time_hr_1 = int(end_time_hr[1]) + 1  # 例如02就取2然后变为整数然后加1就是标签
                                select_time_hr = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[5]/div/div/div/div/div[1]/div/ul[1]/li[{end_time_hr_1}]/div")))  # 锁定结束时间小时标签
                                self.driver.execute_script("arguments[0].click();", select_time_hr)
                            else:
                                end_time_hr_1 = int(end_time_hr) + 1  # 只要不是01-09直接加1就行了，分析标签得出的结论
                                select_time_hr = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[5]/div/div/div/div/div[1]/div/ul[1]/li[{end_time_hr_1}]/div")))  # 锁定结束时间小时标签
                                self.driver.execute_script("arguments[0].click();", select_time_hr)
                            if end_time_minute in list_1:
                                end_time_minute_1 = int(end_time_minute[1]) + 1  # 例如02就取2然后变为整数然后加1就是标签
                                select_time_minute = WebDriverWait(self.driver, 100).until(
                                    ec.presence_of_element_located((By.XPATH, f"/html/body/div[5]/div/div/div/div/div[1]/div/ul[2]/li[{end_time_minute_1}]/div")))  # 锁定结束时间分钟标签 self.driver.execute_script("arguments[0].click();", select_time_minute)
                            else:
                                end_time_minute_1 = int(end_time_minute) + 1  # 只要不是01-09直接加1就行了，分析标签得出的结论
                                select_time_minute = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[5]/div/div/div/div/div[1]/div/ul[2]/li[{end_time_minute_1}]/div")))  # 锁定结束时间分钟标签
                                self.driver.execute_script("arguments[0].click();", select_time_minute)
                            end_time.send_keys(Keys.ENTER)
                            submit_button = WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.XPATH, "/html/body/div[1]/section/section/section[3]/div/div[2]/button[3]/span")))
                            submit_button.click()#点击submit
                            time.sleep(2)
                            self.driver.close()
                            self.driver.switch_to.window(main_window)#打完一份case必须回到主窗口才能继续之前的操作，因为for循环跳到了弹出的窗口
                            time.sleep(1)
                        self.testplan_sheet.cell(column=12, row=every_row, value="执行完成")
                        self.excel.save(r'C:excel_to_TCM_face_to _all_newTCM.xlsx')
                    else:#就是没搜到的情况
                        self.testplan_sheet.cell(column=12, row=every_row, value="TCM无法搜到此case")
                        self.excel.save(r'C:excel_to_TCM_face_to _all_newTCM.xlsx')
                except:
                    try:#如果搜索case成功并且开始打后出错，就尝试先关掉pop up window,再转到主窗口，如是case直接没搜索到出错，这个try就会因为没有pop up window而跳过这步
                        self.driver.switch_to.window(popup_windows)
                        self.driver.close()
                        self.testplan_sheet.cell(column=12, row=every_row,value="回填过程中遇到错误")
                        self.excel.save(r'C:excel_to_TCM_face_to _all_newTCM.xlsx')
                        self.driver.switch_to.window(main_window)
                    except:#眺到这说明没有pop up窗口，是直接没搜索到case。所以不用再用self.driver.switch_to.window(main_window)
                        self.testplan_sheet.cell(column=12, row=every_row, value="case不是not start状态或plan没拉这份case")
                        self.excel.save(r'C:excel_to_TCM_face_to _all_newTCM.xlsx')
                    continue
            else: #very_row_list[10]就不是none就代表有issue
                search_case = self.driver.find_element(By.ID, "rc_select_10")  # 找到搜查case的输入框
                search_case.click()
                search_case.send_keys(Keys.CONTROL, 'a') #全选ctrl+A
                search_case.send_keys(Keys.BACKSPACE)
                time.sleep(1)
                search_case.send_keys(every_row_list[0]) #输入caseID
                search_case.send_keys(Keys.ENTER)
                self.driver.find_element(By.XPATH,"/html/body/div[1]/section/section/section/div/div[2]/div[1]/div/div[2]/div/div[16]/div[2]/div/button[2]/span").click()  # 点击搜索
                time.sleep(3)  # 这里3秒是为了等待显示搜索后的页面，否则后面的代码会找到所有的test cases
                try:#我就默认搜索case后至少有1支case出现
                    all_cases=self.driver.find_elements(By.XPATH, "//*[@id='root']/section/section/section/div/div[2]/div[3]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr")#找到搜索后出现的所有case，这里第一个tr是头标，需要去除
                    len_all_cases=len(all_cases)
                    first_case = self.driver.find_element(By.XPATH, "/html/body/div[1]/section/section/section/div/div[2]/div[3]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr[2]/td[3]/div")  # 锁定正常情况下搜到case的情况，至少有一支case的情况，第4个td标签case ID，如果这里报错，代表搜到了case,但是没有not run的，plan没拉或者被打过
                    if first_case.text == every_row_list[0]:  #如果到了这步，代表上面一步pass了，说明有搜到case并至少出现一份或者直接没有搜到这份case，依然会有上一步的标签, 就是默认没搜索前的第一份case
                        # 如果相等代表是搜到的情况
                        for every_case in range(2,len_all_cases+1):#，遍历所有case,除去第一个tr，因为是标头，所以从2开始，+1是因为range(2,all_cases）只能取到2到all_cases-1,不能遍历所有搜索到的case
                            every_case=self.driver.find_element(By.XPATH, f"/html/body/div[1]/section/section/section/div/div[2]/div[3]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr[{every_case}]/td[18]")
                            self.driver.execute_script("arguments[0].scrollIntoView();", every_case)  # 移到case 元素可见
                            every_case.click()#点击执行case
                            popup_windows = self.driver.window_handles[-1]  # 跳转到新页面执行准备case
                            self.driver.switch_to.window(popup_windows)  #跳转到新页面执行准备case
                            edit_button=WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.XPATH,"/html/body/div/section/section/div[1]/div[2]/button[1]/span")))
                            edit_button.click()
                            fail_button=WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.XPATH,"/html/body/div/section/section/section[3]/div/div[1]/section/div/div[1]/div/div[1]/div/label[3]/span[1]/input")))
                            self.driver.execute_script("arguments[0].click();",fail_button)
                            issue_0_button=WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH,"/html/body/div/section/section/section[3]/div/div[1]/section/div/div[1]/div/div[1]/div/div/span")))
                            self.driver.execute_script("arguments[0].click();", issue_0_button)
                            add_issue_button_first=WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.XPATH,"/html/body/div[2]/div/div[2]/div/div[2]/div[2]/div")))
                            self.driver.execute_script("arguments[0].click();", add_issue_button_first)
                            DFT_list = every_row_list[10].split('\n')#把issue按空格转为为列表
                            issue_id_title_label = 2  # 因为要add多条issue，input标签有变化，第一条issue是2，所以这里设置初始为2，后面观察只需要依次加1就好了
                            if len(DFT_list) == 1:#如果只有1条issue
                                issue_id_input=WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH,"/html/body/div[2]/div/div[2]/div/div[2]/div[2]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr[2]/td[1]/textarea")))
                                issue_id_input.send_keys(DFT_list[0])
                                issue_tittle_input=WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH,"/html/body/div[2]/div/div[2]/div/div[2]/div[2]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr[2]/td[2]/div/textarea")))
                                issue_tittle_input.send_keys(DFT_list[0])
                                self.driver.find_element(By.XPATH,"/html/body/div[2]/div/div[2]/div/div[2]/div[3]/button[2]/span").click()#添加完1条issue后点击add
                            else:#不止1条issue,先添加1条issue后，后面的issue循环添加
                                issue_id_input=WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH,"/html/body/div[2]/div/div[2]/div/div[2]/div[2]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr[2]/td[1]/textarea")))
                                issue_id_input.send_keys(DFT_list[0])
                                issue_tittle_input=WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH,"/html/body/div[2]/div/div[2]/div/div[2]/div[2]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr[2]/td[2]/div/textarea")))
                                issue_tittle_input.send_keys(DFT_list[0])
                                for i in range (1,len(DFT_list)):#从第二条开始提添加:
                                    issue_id_title_label = issue_id_title_label + 1
                                    self.driver.find_element(By.XPATH,"/html/body/div[2]/div/div[2]/div/div[2]/div[2]/div").click()#添加完1条issue后点击add issue准备增加第二条
                                    issue_id_input = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH,f"/html/body/div[2]/div/div[2]/div/div[2]/div[2]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr[{issue_id_title_label}]/td[1]/textarea")))
                                    issue_id_input.send_keys(DFT_list[i])
                                    issue_tittle_input = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH,f"/html/body/div[2]/div/div[2]/div/div[2]/div[2]/section/div/div/div/div/div/div/div/div[2]/table/tbody/tr[{issue_id_title_label}]/td[2]/div/textarea")))
                                    issue_tittle_input.send_keys(DFT_list[i])
                                self.driver.find_element(By.XPATH,"/html/body/div[2]/div/div[2]/div/div[2]/div[3]/button[2]/span").click()  # 添加完所有issue后点击add
                            config=self.driver.find_element(By.XPATH, "//*[@id='root']/section/section/section[2]/div/div/div/div/div/div/div/div/div/table/tbody/tr/td[6]/div")#获取config 标签
                            config_text=config.text#获取config 标签的text值
                            #self.driver.find_element(By.XPATH,"/html/body/div/section/section/section[3]/div/div[1]/div[3]/section[1]/div/div[1]/div/input").send_keys(config_text)#Actual Config标签填值
                            self.driver.find_element(By.XPATH,"/html/body/div/section/section/section[3]/div/div[1]/div[4]/section[1]/div/div[2]/div/input").send_keys(config_text)#Actual DUT标签填值
                            start_date=self.driver.find_element(By.XPATH,"/html/body/div/section/section/section[3]/div/div[1]/div[4]/section[1]/div/div[3]/div/div/div[1]/div/input")#找到开始日期
                            start_date.send_keys(every_row_list[2])#填写开始日期
                            start_time=self.driver.find_element(By.XPATH,"/html/body/div[1]/section/section/section[3]/div/div[1]/div[4]/section[1]/div/div[3]/div/div/div[2]/div/input")#找到开始时间
                            start_time.send_keys(every_row_list[4])#
                            start_time_hr = every_row_list[4][0:2]  # 取开始时间的前2位小时部分
                            start_time_minute = every_row_list[4][3:5]  # 取开始时间的后2位分钟部分
                            if start_time_hr in list_1:
                                start_time_hr_1 = int(start_time_hr[1]) + 1  # 例如02就取2然后变为整数然后加1就是标签
                                select_time_hr = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[3]/div/div/div/div/div[1]/div/ul[1]/li[{start_time_hr_1}]/div")))  # 锁定开始时间小时标签
                                self.driver.execute_script("arguments[0].click();", select_time_hr)
                            else:
                                start_time_hr_1 = int(start_time_hr) + 1  # 只要不是01-09直接加1就行了，分析标签得出的结论
                                select_time_hr = WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[3]/div/div/div/div/div[1]/div/ul[1]/li[{start_time_hr_1}]/div")))  # 锁定开始时间小时标签
                                self.driver.execute_script("arguments[0].click();", select_time_hr)
                            if start_time_minute in list_1:
                                start_time_minute_1 = int(start_time_minute[1]) + 1  # 例如02就取2然后变为整数然后加1就是标签
                                select_time_minute = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[3]/div/div/div/div/div[1]/div/ul[2]/li[{start_time_minute_1}]/div")))  # 锁定开始时间分钟标签
                                self.driver.execute_script("arguments[0].click();", select_time_minute)
                            else:
                                start_time_minute_1 = int(start_time_minute) + 1  # 只要不是01-09直接加1就行了，分析标签得出的结论
                                select_time_minute = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[3]/div/div/div/div/div[1]/div/ul[2]/li[{start_time_minute_1}]/div")))  # 锁定开始时间分钟标签
                                self.driver.execute_script("arguments[0].click();", select_time_minute)
                            start_time.send_keys(Keys.ENTER)
                            end_date=self.driver.find_element(By.XPATH,"/html/body/div[1]/section/section/section[3]/div/div[1]/div[4]/section[1]/div/div[4]/div/div/div[1]/div/input")#找结束日期
                            end_date.click()
                            end_date.send_keys(Keys.CONTROL, 'a')
                            end_date.send_keys(Keys.BACKSPACE)#填写开始日期后默认日期默认填了开始日期，所有先去除默认日期
                            end_date.send_keys(every_row_list[3])#填写结束日期
                            end_date.send_keys(Keys.ENTER)
                            end_time=self.driver.find_element(By.XPATH,"/html/body/div[1]/section/section/section[3]/div/div[1]/div[4]/section[1]/div/div[4]/div/div/div[2]/div/input")#找结束时间
                            end_time.send_keys(every_row_list[5])#填写结束时间
                            end_time_hr = every_row_list[5][0:2]  # 取结束时间的前2位小时部分
                            end_time_minute = every_row_list[5][3:5]
                            if end_time_hr in list_1:
                                end_time_hr_1 = int(end_time_hr[1]) + 1  # 例如02就取2然后变为整数然后加1就是标签
                                select_time_hr = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[5]/div/div/div/div/div[1]/div/ul[1]/li[{end_time_hr_1}]/div")))  # 锁定结束时间小时标签
                                self.driver.execute_script("arguments[0].click();", select_time_hr)
                            else:
                                end_time_hr_1 = int(end_time_hr) + 1  # 只要不是01-09直接加1就行了，分析标签得出的结论
                                select_time_hr = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[5]/div/div/div/div/div[1]/div/ul[1]/li[{end_time_hr_1}]/div")))  # 锁定结束时间小时标签
                                self.driver.execute_script("arguments[0].click();", select_time_hr)
                            if end_time_minute in list_1:
                                end_time_minute_1 = int(end_time_minute[1]) + 1  # 例如02就取2然后变为整数然后加1就是标签
                                select_time_minute = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[5]/div/div/div/div/div[1]/div/ul[2]/li[{end_time_minute_1}]/div")))  # 锁定结束时间分钟标签
                                self.driver.execute_script("arguments[0].click();", select_time_minute)
                            else:
                                end_time_minute_1 = int(end_time_minute) + 1  # 只要不是01-09直接加1就行了，分析标签得出的结论
                                select_time_minute = WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH, f"/html/body/div[5]/div/div/div/div/div[1]/div/ul[2]/li[{end_time_minute_1}]/div")))  # 锁定结束时间分钟标签
                                self.driver.execute_script("arguments[0].click();", select_time_minute)
                            end_time.send_keys(Keys.ENTER)
                            # add_reson_button_first=self.driver.find_element(By.XPATH,"/html/body/div/section/section/section[3]/div/div[1]/div[3]/section[1]/div/div[5]/div/div")
                            # add_reson_button_first.click()
                            # time.sleep(2)
                            # add_reson_second = self.driver.find_element(By.XPATH,"/html/body/div[6]/div/div[2]/div/div[2]/div[2]/div[3]/span").click()
                            # reson_input=WebDriverWait(self.driver, 100).until(ec.presence_of_element_located((By.XPATH,"/html/body/div[6]/div/div[2]/div/div[2]/div[2]/div[2]/div/div[1]/div[1]/div/span[1]/input")))
                            # self.driver.execute_script("arguments[0].removeAttribute('readonly');", reson_input)  # 移除只读属性
                            # reson_input.send_keys(every_row_list[6])
                            # time.sleep(1)
                            # reson_input.send_keys(Keys.ENTER)
                            # time.sleep(1)
                            # Hrs = self.driver.find_element(By.XPATH,"/html/body/div[6]/div/div[2]/div/div[2]/div[2]/div[2]/div/div[1]/div[2]/div[2]/input")
                            # Hrs.send_keys('1')  # 先随便填个non  test 时间
                            # self.driver.find_element(By.XPATH,"/html/body/div[6]/div/div[2]/div/div[2]/div[2]/div[2]/div/div[2]/div[2]/textarea").click()  # 需要点一下comment再重新填non test时间
                            # Hrs.send_keys(Keys.BACKSPACE)  # 点了comment后之前的Hrs栏位会变为0.0，所以按三下backspace清掉
                            # Hrs.send_keys(Keys.BACKSPACE)
                            # Hrs.send_keys(Keys.BACKSPACE)
                            # Hrs.send_keys(every_row_list[7])
                            # Hrs.find_element(By.XPATH,"/html/body/div[6]/div/div[2]/div/div[2]/div[2]/div[2]/div/div[2]/div[2]/textarea").click()  # 再点一下comment看看能不能解决reson有时候保存不上的问题
                            # self.driver.find_element(By.XPATH,"/html/body/div[6]/div/div[2]/div/div[2]/div[3]/button[2]/span").click()  # 点击add保存
                            # time.sleep
                            submit_button=WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.XPATH,"/html/body/div[1]/section/section/section[3]/div/div[2]/button[3]/span")))
                            submit_button.click()#点击submit
                            time.sleep(2)
                            self.driver.close()
                            self.driver.switch_to.window(main_window)#打完一份case必须回到主窗口才能继续之前的操作，因为for循环跳到了弹出的窗口
                            time.sleep(1)
                        self.testplan_sheet.cell(column=12, row=every_row, value="执行完成")
                        self.excel.save(r'C:excel_to_TCM_face_to _all_newTCM.xlsx')
                    else:# 就是没搜到的情况
                        self.testplan_sheet.cell(column=12, row=every_row, value="TCM无法搜到此case")
                        self.excel.save(r'C:excel_to_TCM_face_to _all_newTCM.xlsx')
                except:
                    try:  # 如果搜索case成功并且开始打后出错，就尝试先关掉pop up window,再转到主窗口，如是case直接没搜索到出错，这个try就会因为没有pop up window而跳过这步
                        self.driver.switch_to.window(popup_windows)
                        self.driver.close()
                        self.testplan_sheet.cell(column=12, row=every_row, value="回填过程中遇到错误")
                        self.excel.save(r'C:excel_to_TCM_face_to _all_newTCM.xlsx')
                        self.driver.switch_to.window(main_window)
                    except:#眺到这说明没有pop up窗口，是直接没搜索到case。所以不用再用self.driver.switch_to.window(main_window)
                        self.testplan_sheet.cell(column=12, row=every_row,value="case不是not start状态或plan没拉这份case")
                        self.excel.save(r'C:excel_to_TCM_face_to _all_newTCM.xlsx')
                    continue







    def run(self):
        self.enter_not_start_case()
        self.enter_case()


if __name__ == "__main__":
        case_go = TCM()
        case_go.run()

